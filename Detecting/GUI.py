from PyQt5 import QtWidgets
from PyQt5 import QtCore
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox, QFrame, QListWidgetItem, QLabel, QSizePolicy
from PyQt5.QtGui import QPixmap, QIntValidator
from PyQt5.QtCore import Qt, QRect

from Detection import Detector, ModelNotValidException, JSONNotValidException
from Logs import JLogs

import os
import math

from Utils import *

from Immagini import *

from Stack import Stack

from openpyxl import Workbook
from openpyxl.styles import Font

class Verifica(QFrame):
    
    def __init__(self, home):
        
        super(Verifica, self).__init__()

        loadUi("guis//verifica.ui", self)
        
        self.home = home
        
        self.stack = Stack()
        
        self.objIndex = -1
        
        QApplication.instance().focusChanged.connect(self.__onFocusChanged)
        
        self.sons = self.__inizializzaJSON()
        
        self.__inizializzaTreeFiles()
        
        self.__inizializzaBottoni()
        
        self.listFiles.keyPressEvent = self.keyPressEvent
        
        self.listFiles.setFocus(True)
    
    # Inizializzazioni
    def __inizializzaBottoni(self):
    
        self.listFiles.clicked.connect(self.__clicked)
        
        self.btnImgSuccessivo.clicked.connect(self.__prossimaImmagine)
        
        self.btnImgPrecedente.clicked.connect(self.__immaginePrecedente)
        
        self.btnCropSuccessivo.clicked.connect(self.__prossimoCrop)
        
        self.btnCropPrecedente.clicked.connect(self.__cropPrecedente)
        
        self.btnUndo.clicked.connect(self.__undo)
        
        self.btnSalva.clicked.connect(lambda: self.__salva(item = self.listFiles.item(self.listFiles.currentRow())))
        
        self.btnEliminaImmagine.clicked.connect(lambda: self.__eliminaImmagine(item = self.listFiles.item(self.listFiles.currentRow())))
        
        self.btnEliminaCrop.clicked.connect(lambda: self.__eliminaCrop(item = self.listFiles.item(self.listFiles.currentRow())))
        
        self.btnEsporta.clicked.connect(self.__esporta)
        
        self.textX.setValidator(QIntValidator())
        
        self.textY.setValidator(QIntValidator())
        
        self.textLarghezza.setValidator(QIntValidator())
        
        self.textLunghezza.setValidator(QIntValidator())
        
    def __inizializzaJSON(self):
        
        try:
        
            return leggiJSON(path = self.home.textFine.toPlainText())
            
        except FileNotFoundError as ex:
            
            creaAvviso(parent = self, messaggio = 'JSON non trovato', icona = QMessageBox.Critical, func = self.chiudi)
            
            return None
        
        except JSONNotValidException as ex:
        
            creaAvviso(parent = self, messaggio = 'File JSON vuoto!', icona = QMessageBox.Critical, func = self.chiudi)
            
            return None
    
    def __inizializzaTreeFiles(self):
        
        if self.sons is None: 
            
            return
        
        else: 
        
            self.showMaximized()
        
        for file in self.sons.keys():

            try:
        
                if fileValido(file) and len(self.sons[file]['objects']):
                    
                    self.listFiles.addItem(QListWidgetItem(file))
            
            except KeyError as ex:
                
                self.home.logs2.append(f'Nel file JSON non sono presenti le informazioni di "{os.path.basename(file)}"')
                
        self.listFiles.setCurrentRow(0)
        
        self.__cambiaImmagine(item = self.listFiles.item(self.listFiles.currentRow()))
    
    def __inizializzaExcel(self, sheet):
        
        for column, text in enumerate(['', 'PATH', 'RUN', 'FILENAME', 'CARTELLO', 'ACCURATEZZA', 'COORDINATE'], start = 1):
            
            sheet.cell(row = 1, column = column, value = text).font = Font(bold = 1)

    # Metodi immagine originale
    def __clicked(self):
        
        self.__cambiaImmagine(item = self.listFiles.item(self.listFiles.currentRow()))
        
    def __setImmagineOriginale(self, item):

        self.listFiles.setCurrentItem(item)

        self.labelTitolo.setText(os.path.basename(item.text()))
        
        self.indiceImg.setText(f'{self.listFiles.currentRow() + 1} di {self.listFiles.count()}')
        
        self.img.setPixmap(QPixmap(self.listFiles.item(self.listFiles.currentRow()).text()))
    
    def __cambiaImmagine(self, item):
        
        if item is None: return 
        
        self.objIndex = -1
        
        self.__setImmagineOriginale(item = item)
        
        self.__cropImg(item = item)
        
        self.__setInfoImmagine(item = item)
        
    def __prossimaImmagine(self):
        
        self.__cambiaImmagine(item = self.listFiles.item(self.listFiles.currentRow() + 1))
        
    def __immaginePrecedente(self):
        
        self.__cambiaImmagine(item = self.listFiles.item(self.listFiles.currentRow() - 1))
    
    # Metodo per mostrare le informazioni trovate
    def __setInfoImmagine(self, item, incremento = 1):
        
        # Se l'indice corrente sommato all'incremento è nel range consentito possiamo effettuare l'incremento senza problemi,
        # alterimenti abbiamo già raggiunto il massimo/minimo consentiti
        if -1 < self.objIndex + incremento < len(self.sons[item.text()]['objects']):
            
            self.objIndex += incremento
            
        self.stack.clear()
            
        self.textPath.setText(self.sons[item.text()]['path'])
        
        self.textRun.setText(self.sons[item.text()]['run'])
        
        self.textFilename.setText(self.sons[item.text()]['filename'])
                
        self.textTipo.setText(self.sons[item.text()]['objects'][self.objIndex]['name'])

        self.textPercentuale.setText(str(self.sons[item.text()]['objects'][self.objIndex]['percentage_probability']))

        self.textX.setText(str(self.sons[item.text()]['objects'][self.objIndex]['box_points'][0]))

        self.textY.setText(str(self.sons[item.text()]['objects'][self.objIndex]['box_points'][1]))

        self.textLarghezza.setText(str(self.sons[item.text()]['objects'][self.objIndex]['box_points'][2]))

        self.textLunghezza.setText(str(self.sons[item.text()]['objects'][self.objIndex]['box_points'][3]))   
        
        self.__setPreview(item = item)       
        
        QApplication.processEvents()
    
    # Metodi per mostrare gli oggetti trovati nella immagine originale
    def __prossimoCrop(self):
        
        self.__setInfoImmagine(item = self.listFiles.item(self.listFiles.currentRow()), incremento = 1)
        
    def __cropPrecedente(self):
        
        self.__setInfoImmagine(item = self.listFiles.item(self.listFiles.currentRow()), incremento = -1)
        
    def __cropImg(self, item):
        
        self.__clearCrops()

        objects = self.sons[item.text()]['objects']

        # Quanti riquadri devo creare per il gridLayout in base agli elementi che ho. Calcolo: Radice (perfetta) di len(oggetti). Quindi 2x2 per 3/4 elemnti, 3x3 per 5 < elemnti < 10, ecc...
        n = math.floor(math.sqrt(len(objects))) + 1 if len(objects) > 2 else len(objects)
        index = 0
        
        for i in range(n):
            
            for j in range(n):
                
                try:
                    
                    x, y, x1, y1 = objects[index]['box_points']
                    
                    label = QLabel(self)
                    
                    label.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Maximum)
                    
                    label.setAlignment(Qt.AlignCenter)
                    
                    label.setMaximumSize(250 / n + 10, 250 / n + 10)
                    
                    img = QPixmap(item.text()).copy(QRect(x, y, x1 - x, y1 - y))
                    
                    label.setPixmap(img)
                    
                    self.gridLayout.addWidget(label, i, j)
                    
                    index += 1
                    
                except IndexError as ex:
                    
                    return 
        
    def __clearCrops(self):
        
        while self.gridLayout.count():
            
            self.gridLayout.takeAt(0).widget().deleteLater()
    
    # Metodo per mostrare l'oggetto a cui si riferiscono le immagini
    def __setPreview(self, item):
        
        x, y, x1, y1 = self.sons[item.text()]['objects'][self.objIndex]['box_points']
        
        img = QPixmap(item.text()).copy(QRect(x, y, x1 - x, y1 - y)).scaled(150, 150)

        self.imgPreview.setPixmap(img)
    
    # Eventi da tastiera    
    def keyPressEvent(self, event):
        
        if event.key() == QtCore.Qt.Key_D:
            
            self.__prossimaImmagine()
        
        elif event.key() == QtCore.Qt.Key_A:
        
            self.__immaginePrecedente()
        
        elif event.key() == QtCore.Qt.Key_Right:
            
            self.__setInfoImmagine(item = self.listFiles.item(self.listFiles.currentRow()))
            
        elif event.key() == QtCore.Qt.Key_Left:
            
            self.__setInfoImmagine(item = self.listFiles.item(self.listFiles.currentRow()), incremento = -1)
        
        else:
            
            super().keyPressEvent(event)
    
    def __salva(self, item):

        self.sons[item.text()]['objects'][self.objIndex] = {
            'name': self.textTipo.text(),
            'percentage_probability': self.textPercentuale.text(),
            'box_points': [int(coordinata) for coordinata in [self.textX.text(), self.textY.text(), self.textLarghezza.text(), self.textLunghezza.text()]] # Prendiamo le coordinate e le convertiamo in interi
        }
        
        salvaSon(son = self.sons[item.text()], output = toJSON(file = item.text()))
    
        salvaSons(sons = self.sons, output = toJSON(file = self.home.textFine.toPlainText()))
    
    def __onFocusChanged(self):
        
        # Se clicco un campo di testo lo appendo allo stack (questo mi dirà che è l'ultimo modificato)
        if QApplication.focusWidget() in (self.textPath, self.textRun, self.textFilename, self.textTipo, self.textPercentuale, self.textX, self.textY, self.textLarghezza, self.textLunghezza):
            
            self.stack.append(QApplication.focusWidget())
     
    def __undo(self):
        
        # Prima di rimuovere l'ultimo elemento modificato effettuiamo l'undo
        if not self.stack.isEmpty():
            
            self.stack.pop().undo()
          
    def __esporta(self):  # sourcery skip: comprehension-to-generator
        
        workbook = Workbook()
        
        sheet = workbook.active
        
        self.__inizializzaExcel(sheet = sheet)

        # Inizia da 1 e non 0
        row = 2
        
        for i in range(self.listFiles.count()):
            
            key = self.listFiles.item(i).text()
            
            # Controllo solo le immagini con oggetti che trovo
            if len(self.sons[key]['objects']) > 0:
                
                sheet.cell(row = row, column = 1, value = i + 1)
                
                sheet.cell(row = row, column = 2, value = self.sons[key]['path'])

                sheet.cell(row = row, column = 3, value = self.sons[key]['run'])

                sheet.cell(row = row, column = 4, value = self.sons[key]['filename'])

                self.__controllaLarghezzaCella(sheet = sheet, data = self.sons[key]['filename'], column = 'D')

                # Ogni oggetto avrà la sua riga
                for obj in self.sons[key]['objects']:

                    sheet.cell(row = row, column = 5, value = obj['name'])
                    
                    self.__controllaLarghezzaCella(sheet = sheet, data = obj['name'], column = 'E')
                    
                    sheet.cell(row = row, column = 6, value = obj['percentage_probability']).number_format = '##.###'
                    
                    self.__controllaLarghezzaCella(sheet = sheet, data = str(obj['percentage_probability']), column = 'F')
                    
                    sheet.cell(row = row, column = 7, value = ' '.join([str(i) for i in obj['box_points']]))
                    
                    self.__controllaLarghezzaCella(sheet = sheet, data = ' '.join([str(i) for i in obj['box_points']]), column = 'G')
                    
                    row += 1
                    
        
        try:
                
            workbook.save(filename = toXLSX(file = self.home.textFine.toPlainText()))

        except PermissionError:
                
            creaAvviso(self, messaggio = 'Chiudi il file in excel prima!', icona = QMessageBox.Information, func = lambda: None)

    def __controllaLarghezzaCella(self, sheet, data, column, offset = 1.2):
        
        if len(data) > sheet.column_dimensions[column].width:

            sheet.column_dimensions[column].width = len(data) * offset
                        
    def __eliminaImmagine(self, item):
        
        del self.sons[item.text()]
        
        self.listFiles.takeItem(self.listFiles.currentRow())
        
        QApplication.processEvents()
        
        salvaSons(sons = self.sons, output = toJSON(file = self.home.textFine.toPlainText()))
        
        self.__cambiaImmagine(item = self.listFiles.item(self.listFiles.currentRow() + 1))
    
    def __eliminaCrop(self, item):  # sourcery skip: merge-else-if-into-elif
        
        del self.sons[item.text()]['objects'][self.objIndex]
        
        if len(self.sons[item.text()]['objects']) == 0:
            
            self.__eliminaImmagine(item = item)
            
        else:
            
            if -1 < self.objIndex + 1 < len(self.sons[item.text()]['objects']):
                
                self.objIndex += 1
                
            else:
                
                self.objIndex -= 1
                
            self.__cambiaImmagine(item = self.listFiles.item(self.listFiles.currentRow()))
        
    def closeEvent(self, event):
        
        self.chiudi()
       
    def chiudi(self):
        
        self.hide()
        
        self.home.show()
    
class Home(QMainWindow):

    def __init__(self):
        
        super(Home, self).__init__()
        
        loadUi("guis//home.ui", self)

        self.setFixedSize(1280, 700)
        
        self.__inizializzaBottoni()
        
        self.files = None
        
        self.running = False
        
        self.detector = Detector()
        
        self.jlogs = JLogs()
        
        self.index = 0

    # Inizializzazioni
    def __inizializzaBottoni(self):
        
        self.btnModello.clicked.connect(self.__selezionaModello)
        
        self.btnJSON.clicked.connect(self.__selezionaJSON)
        
        self.btnInizio.clicked.connect(self.__selezionaInizio)
        
        self.btnFine.clicked.connect(self.__selezionaFine)
        
        self.btnStart.clicked.connect(self.start)
        
        self.btnStop.clicked.connect(self.stop)
        
        self.btnApriFoto.clicked.connect(lambda: os.startfile(self.textInizio.toPlainText()))
        
        self.btnApriSalvataggi.clicked.connect(lambda: os.startfile(self.textFine.toPlainText()))
        
        self.spinBox.valueChanged.connect(self.__percentualeCambiata)
        
        self.btnVerifica.clicked.connect(self.__verifica)
        
    def __inizializzaProgressBar(self):
        
        self.progressBar.reset()
        
        self.spinBox.setValue(self.detector.PERCENTUALE_MINIMA)
        
        # Prendiamo i file della cartella di destinazione
        self.files = [os.path.join(path, name) for path, subdirs, files in os.walk(self.textInizio.toPlainText()) for name in files]
        
        self.progressBar.setRange(0, len(self.files))
    
    def __inizializzaDetector(self):
        
        self.logs1.append('Inizializzando Tensorflow...')
        
        QApplication.processEvents()
        
        try:
        
            self.detector.setModelPath(self.textModello.toPlainText())
            
            self.detector.setJsonPath(self.textJSON.toPlainText())
            
            self.detector.loadModel()
            
        except ModelNotValidException:
    
            creaAvviso(parent = self, messaggio = 'Seleziona un modello valido!', icona = QMessageBox.Critical, func = lambda: self.logs2.append('Modello non valido'))
        
        except JSONNotValidException:
            
            creaAvviso(parent = self, messaggio = 'Seleziona un JSON valido!', icona = QMessageBox.Critical, func = lambda: self.logs2.append('JSON non valido'))
    
    # Metodi per selezionare i diversi file/cartelle
    def __selezionaModello(self):
        
        self.index = 0
        
        self.__openFileNameDialog(caption = 'Seleziona modello', files = "H5 Files (*.h5)", textField = self.textModello)
        
    def __selezionaJSON(self):
        
        self.index = 0
        
        self.__openFileNameDialog(caption = 'Seleziona JSON', files = "JSON File (*.json)", textField = self.textJSON)
        
    def __selezionaInizio(self):
        
        self.index = 0
        
        self.__openFolder(caption = 'Seleziona cartella foto', textField = self.textInizio)
        
        self.__inizializzaProgressBar()
        
    def __selezionaFine(self):
        
        self.__openFolder(caption = 'Seleziona cartella salvataggi', textField = self.textFine)

    def __percentualeCambiata(self):
        
        self.detector.PERCENTUALE_MINIMA = self.spinBox.value()

    """Metodo che mostra la seconda finestra se la cartella di destinazione è stata scelta"""
    def __verifica(self):

        if self.textFine.toPlainText() == '':
            
            creaAvviso(parent = self, messaggio = 'Seleziona una cartella di salvataggi valida!', icona = QMessageBox.Critical, func = lambda: self.logs2.append('Cartella salvataggi non valida'))
            
        else:

            self.hide()
            
            self.verifica = Verifica(self)
            
    # File/Folder Dialogs
    def __openFolder(self, caption, textField):
        
        fileName = QFileDialog.getExistingDirectory(self, caption = caption, directory = textField.toPlainText())
        
        if fileName:
            
            textField.setText(fileName)

    def __openFileNameDialog(self, caption, files, textField):
        
        fileName, _ = QFileDialog.getOpenFileName(self, caption = caption, directory = textField.toPlainText(), filter = files)
        
        if fileName:
            
            textField.setText(fileName)
    
    """Metodo per far partire la detenzione"""
    def start(self):
        
        if not self.detector.same_model(self.textModello.toPlainText()):         
        
            self.__inizializzaDetector()
            
        self.running = True
        
        self.__run()
    
    """Metodo per mettere in pausa la detenzione"""
    def stop(self):
        
        self.running = False
        
    """
    Metodo per scorrere tutte le immagini che abbiamo e salvare le informazioni in un json.
    N.B. Utilizza start() per iniziare la detenzione.
    """
    def __run(self):
        
        self.__inizializzaProgressBar()
        
        while self.running and self.index < len(self.files):
        
            # Controlliamo se il file è supportato
            if fileValido(self.files[self.index]):
        
                self.logs1.append(f'Detecting {self.files[self.index]}...')           
                QApplication.processEvents()
                try:
    
                    # Prendiamo la path del file in cui verrà salvato e controlliamo se esiste
                    save = self.files[self.index].replace(self.textInizio.toPlainText(), self.textFine.toPlainText())
                    
                    folder = os.path.dirname(save)
                    
                    while not os.path.exists(folder):
    
                        self.__creaAlberoFiles(folder)
                    
                    self.jlogs.salva(detections = self.detector.detect(self.files[self.index], save), file = save, output = toJSON(file = save))
                    
                except Exception as e:
                    
                    self.logs2.append(f'Erore {save}: {e}...')    
                
                else:
                
                    self.logs2.append(f'Saving to {save}...')
                
            else:
                
                self.logs1.append(f'{self.files[self.index]} non valido...')
            
            QApplication.processEvents()
            
            self.index += 1
    
            self.progressBar.setValue(self.index)
            
            # Se la barra raggiunge il 100% il file verrà salvato
            if self.progressBar.value() == self.progressBar.maximum():
            
                self.jlogs.salvaLog(output = toJSON(self.textFine.toPlainText()), nomeModello = self.detector.getModelPath(), percentuale = self.detector.PERCENTUALE_MINIMA)
                
                self.running = False
                
                self.index = 0
   
    """Metodo per ricreare l'albero delle cartelle originale per utilizzare la medesima struttura"""
    def __creaAlberoFiles(self, cartella):
    
        if not os.path.exists(cartella):
            
            try:
                
                os.mkdir(cartella)
                
            except FileNotFoundError:
                
                self.creaAlberoFiles(os.path.dirname(cartella))
                
if __name__ == "__main__":
    
    import sys
    app = QtWidgets.QApplication(sys.argv)
    home = Home()
    home.show()
    sys.exit(app.exec_())
